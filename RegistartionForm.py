import tkinter as tk
from tkinter import messagebox
import openpyxl

# Function to handle the registration button click event
def register():
    name = entry_name.get()
    email = entry_email.get()
    phone = entry_phone.get()

    # Validate input fields
    if not name or not email or not phone:
        messagebox.showerror("Error", "Please fill in all the fields.")
        return

    # Open the Excel workbook and select the active sheet
    workbook = openpyxl.load_workbook('registrations.xlsx')
    sheet = workbook.active

    # Find the first empty row in the Excel sheet
    row = sheet.max_row + 1

    # Write the registration details to the Excel sheet
    sheet.cell(row=row, column=1).value = name
    sheet.cell(row=row, column=2).value = email
    sheet.cell(row=row, column=3).value = phone

    # Save the workbook
    workbook.save('registrations.xlsx')

    # Display success message
    messagebox.showinfo("Success", "Registration successful.")

    # Clear the input fields
    entry_name.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_phone.delete(0, tk.END)

# Function to display all registered details in a new window
def display_registered_details():
    # Open the Excel workbook and select the active sheet
    workbook = openpyxl.load_workbook('registrations.xlsx')
    sheet = workbook.active

    # Get the total number of rows in the Excel sheet
    num_rows = sheet.max_row

    # Create a new window to display the registered details
    window = tk.Toplevel(root)
    window.title("Registered Details")

    # Create a label for each registered detail
    for row in range(1, num_rows + 1):
        name = sheet.cell(row=row, column=1).value
        email = sheet.cell(row=row, column=2).value
        phone = sheet.cell(row=row, column=3).value
        label = tk.Label(window, text=f"Name: {name}\nEmail: {email}\nPhone: {phone}")
        label.pack()



# Create the main window
root = tk.Tk()
root.title("Registration Form")
root.geometry("200x200")
# Create labels and entry fields for name,email and phone number
label_name = tk.Label(root, text="Name:")
label_name.grid(row=0,column=0)
entry_name = tk.Entry(root)
entry_name.grid(row=0,column=1)

label_email = tk.Label(root, text="Email:")
label_email.grid(row=1,column=0)
entry_email = tk.Entry(root)
entry_email.grid(row=1,column=1)

label_phone = tk.Label(root, text="Phone:")
label_phone.grid(row=2,column=0)
entry_phone = tk.Entry(root)
entry_phone.grid(row=2,column=1)

# Create the registration button
register_button = tk.Button(root, text="Register", command=register)
register_button.grid(row=3,column=0)

# Create the button to display registered details
display_button = tk.Button(root, text="Display Registered Details", 
                           command=display_registered_details)
display_button.grid(row=3,column=1)

# Start the main Tkinter event loop
root.mainloop()
